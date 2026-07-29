#!/usr/bin/env python3
"""
sponsorscan - find entry-level jobs at employers with real H-1B filing history.

Two data sources, both of which are structurally durable:

  1. DOL OFLC LCA disclosure data. Published quarterly by the Office of Foreign
     Labor Certification. A legal filing requirement, so it keeps coming out
     regardless of anyone's business model. Free bulk download, no key.

  2. Public ATS job board APIs (Greenhouse, Lever, Ashby). Served straight from
     the employer with no aggregator in between to abandon it.

Usage:
    python sponsorscan.py load-lca ~/Downloads/LCA_Disclosure_Data_FY2026_Q2.xlsx
    python sponsorscan.py fetch-jobs
    python sponsorscan.py report --out matches.csv

Run `python sponsorscan.py <command> --help` for per-command options.
"""

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
import time
import urllib.parse

import requests
import yaml

try:
    from rapidfuzz import process as rf_process
    from rapidfuzz import fuzz as rf_fuzz
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

DB_PATH = os.environ.get("SPONSORSCAN_DB", "sponsorscan.db")
UA = {"User-Agent": "sponsorscan/1.0 (personal job search tool)"}


# ---------------------------------------------------------------- normalization

# Suffixes and filler tokens stripped before matching employer names. "Google
# LLC", "GOOGLE INC." and "Google, Inc" all collapse to "google".
_SUFFIXES = {
    "inc", "incorporated", "llc", "l l c", "lp", "llp", "plc", "ltd", "limited",
    "co", "corp", "corporation", "company", "holdings", "holding", "group",
    "usa", "us", "america", "americas", "na", "the", "and", "of",
}

_PUNCT = re.compile(r"[^a-z0-9 ]+")
_WS = re.compile(r"\s+")


def norm_employer(name):
    """Collapse an employer name to a comparable key."""
    if not name:
        return ""
    s = str(name).lower()
    s = s.replace("&", " and ")
    s = _PUNCT.sub(" ", s)
    toks = [t for t in _WS.sub(" ", s).strip().split(" ") if t and t not in _SUFFIXES]
    return " ".join(toks)


# ------------------------------------------------------------------- db schema

SCHEMA = """
CREATE TABLE IF NOT EXISTS employers (
    employer_norm    TEXT PRIMARY KEY,
    employer_display TEXT,
    certified        INTEGER DEFAULT 0,
    denied           INTEGER DEFAULT 0,
    withdrawn        INTEGER DEFAULT 0,
    titles           TEXT,
    states           TEXT,
    wage_samples     TEXT,
    lvl1 INTEGER DEFAULT 0,
    lvl2 INTEGER DEFAULT 0,
    lvl3 INTEGER DEFAULT 0,
    lvl4 INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS jobs (
    job_key      TEXT PRIMARY KEY,
    source       TEXT,
    company      TEXT,
    company_norm TEXT,
    title        TEXT,
    location     TEXT,
    url          TEXT,
    posted       TEXT,
    description  TEXT,
    fetched_at   TEXT
);
CREATE INDEX IF NOT EXISTS idx_jobs_norm ON jobs(company_norm);
CREATE TABLE IF NOT EXISTS probe_cache (
    provider   TEXT,
    slug       TEXT,
    ok         INTEGER,
    n_jobs     INTEGER,
    checked_at TEXT,
    PRIMARY KEY (provider, slug)
);
"""


def connect():
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    have = {r[1] for r in con.execute("PRAGMA table_info(employers)")}
    for col in ("lvl1", "lvl2", "lvl3", "lvl4"):
        if col not in have:
            con.execute(f"ALTER TABLE employers ADD COLUMN {col} INTEGER DEFAULT 0")
    con.commit()
    return con


# ------------------------------------------------------------------- load-lca

# The DOL file has ~100 columns and the exact header set drifts between fiscal
# years, so columns are resolved by name rather than position.
WANTED = {
    "employer": ["EMPLOYER_NAME"],
    "status": ["CASE_STATUS"],
    "title": ["JOB_TITLE", "SOC_TITLE"],
    "state": ["WORKSITE_STATE", "WORKSITE_STATE_1", "EMPLOYER_STATE"],
    "wage": ["WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_FROM_1"],
    "wage_unit": ["WAGE_UNIT_OF_PAY", "WAGE_UNIT_OF_PAY_1"],
    # Prevailing wage level (I-IV). Under the FY2027 weighted selection rule,
    # petitions filed at Level III/IV get better lottery odds, so an employer's
    # typical level directly affects your chances, not just your pay.
    "wage_level": ["PW_WAGE_LEVEL", "PW_WAGE_LEVEL_1"],
}


def _resolve_columns(header):
    idx = {}
    upper = [str(h).strip().upper() if h is not None else "" for h in header]
    for key, candidates in WANTED.items():
        for cand in candidates:
            if cand in upper:
                idx[key] = upper.index(cand)
                break
    missing = [k for k in ("employer", "status") if k not in idx]
    if missing:
        raise SystemExit(
            f"Could not find required column(s) {missing} in the file header.\n"
            f"Saw: {upper[:25]}..."
        )
    return idx


def _iter_rows(path):
    """Stream rows from .xlsx or .csv without loading the whole file."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            for row in csv.reader(fh):
                yield row


def _annual_wage(value, unit):
    try:
        v = float(str(value).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return None
    u = (str(unit) or "").strip().lower()
    mult = {"year": 1, "hour": 2080, "week": 52, "bi-weekly": 26,
            "biweekly": 26, "month": 12}.get(u, 1)
    v *= mult
    return v if 10_000 < v < 2_000_000 else None


def cmd_load_lca(args):
    src = args.path
    if src.startswith(("http://", "https://")):
        local = os.path.basename(urllib.parse.urlparse(src).path) or "lca_download.xlsx"
        print(f"Downloading {src} -> {local} (this file is typically 100-400 MB)")
        with requests.get(src, stream=True, headers=UA, timeout=120) as r:
            r.raise_for_status()
            with open(local, "wb") as fh:
                for chunk in r.iter_content(1 << 20):
                    fh.write(chunk)
        src = local

    if not os.path.exists(src):
        raise SystemExit(f"No such file: {src}")

    con = connect()
    if args.replace:
        con.execute("DELETE FROM employers")

    agg = {}
    rows = _iter_rows(src)
    try:
        header = next(rows)
    except StopIteration:
        raise SystemExit("File is empty.")
    idx = _resolve_columns(header)

    n = 0
    for row in rows:
        n += 1
        if n % 100_000 == 0:
            print(f"  ...{n:,} rows")

        def get(key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        emp = get("employer")
        if not emp:
            continue
        key = norm_employer(emp)
        if not key:
            continue

        rec = agg.setdefault(key, {
            "display": str(emp).strip(), "certified": 0, "denied": 0,
            "withdrawn": 0, "titles": {}, "states": {}, "wages": [],
            "lvl": {1: 0, 2: 0, 3: 0, 4: 0},
        })

        status = (str(get("status")) or "").strip().upper()
        if status.startswith("CERTIFIED") and "WITHDRAWN" in status:
            rec["withdrawn"] += 1
        elif status.startswith("CERTIFIED"):
            rec["certified"] += 1
        elif status.startswith("DENIED"):
            rec["denied"] += 1
        elif status.startswith("WITHDRAWN"):
            rec["withdrawn"] += 1

        t = get("title")
        if t:
            t = str(t).strip().lower()[:80]
            rec["titles"][t] = rec["titles"].get(t, 0) + 1
        st = get("state")
        if st:
            st = str(st).strip().upper()[:2]
            rec["states"][st] = rec["states"].get(st, 0) + 1
        lvl = str(get("wage_level") or "").strip().upper().replace("LEVEL", "").strip()
        lvl_n = {"I": 1, "II": 2, "III": 3, "IV": 4,
                 "1": 1, "2": 2, "3": 3, "4": 4}.get(lvl)
        if lvl_n:
            rec["lvl"][lvl_n] += 1

        w = _annual_wage(get("wage"), get("wage_unit"))
        if w and len(rec["wages"]) < 400:
            rec["wages"].append(w)

    print(f"Read {n:,} rows, {len(agg):,} distinct employers.")

    payload = []
    for key, rec in agg.items():
        top_titles = sorted(rec["titles"].items(), key=lambda kv: -kv[1])[:12]
        top_states = sorted(rec["states"].items(), key=lambda kv: -kv[1])[:8]
        payload.append((
            key, rec["display"], rec["certified"], rec["denied"], rec["withdrawn"],
            json.dumps([t for t, _ in top_titles]),
            json.dumps([s for s, _ in top_states]),
            json.dumps(sorted(rec["wages"])[:200]),
            rec["lvl"][1], rec["lvl"][2], rec["lvl"][3], rec["lvl"][4],
        ))

    con.executemany(
        "INSERT OR REPLACE INTO employers "
        "(employer_norm, employer_display, certified, denied, withdrawn, titles, states, "
        " wage_samples, lvl1, lvl2, lvl3, lvl4) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", payload)
    con.commit()
    print(f"Loaded {len(payload):,} employers into {DB_PATH}")


# ------------------------------------------------------------------ fetch-jobs
#
# NOTE: every function in this section makes a live HTTP call and could NOT be
# tested in the environment where this was written. If a provider changes its
# response shape, this is the first place to look.

def _get_json(url, timeout=25):
    r = requests.get(url, headers=UA, timeout=timeout)
    r.raise_for_status()
    return r.json()


def fetch_greenhouse(slug):
    url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true"
    data = _get_json(url)
    out = []
    for j in data.get("jobs", []):
        loc = (j.get("location") or {}).get("name", "")
        out.append({
            "job_key": f"greenhouse:{slug}:{j.get('id')}",
            "source": "greenhouse", "title": j.get("title", ""),
            "location": loc, "url": j.get("absolute_url", ""),
            "posted": (j.get("updated_at") or "")[:10],
            "description": j.get("content", "") or "",
        })
    return out


def fetch_lever(slug):
    url = f"https://api.lever.co/v0/postings/{slug}?mode=json"
    data = _get_json(url)
    out = []
    for j in data:
        cats = j.get("categories") or {}
        out.append({
            "job_key": f"lever:{slug}:{j.get('id')}",
            "source": "lever", "title": j.get("text", ""),
            "location": cats.get("location", "") or "",
            "url": j.get("hostedUrl", ""),
            "posted": time.strftime("%Y-%m-%d", time.gmtime((j.get("createdAt") or 0) / 1000))
                      if j.get("createdAt") else "",
            "description": (j.get("descriptionPlain") or "") +
                           " " + json.dumps(j.get("lists", [])),
        })
    return out


def fetch_ashby(slug):
    url = f"https://api.ashbyhq.com/posting-api/job-board/{slug}?includeCompensation=true"
    data = _get_json(url)
    out = []
    for j in data.get("jobs", []):
        out.append({
            "job_key": f"ashby:{slug}:{j.get('id')}",
            "source": "ashby", "title": j.get("title", ""),
            "location": j.get("location", "") or "",
            "url": j.get("jobUrl", "") or j.get("applyUrl", ""),
            "posted": (j.get("publishedAt") or "")[:10],
            "description": j.get("descriptionPlain", "") or "",
        })
    return out


FETCHERS = {"greenhouse": fetch_greenhouse, "lever": fetch_lever, "ashby": fetch_ashby}


def cmd_fetch_jobs(args):
    with open(args.companies) as fh:
        cfg = yaml.safe_load(fh)

    con = connect()
    if args.replace:
        con.execute("DELETE FROM jobs")

    total, failed = 0, []
    for provider, entries in (cfg.get("companies") or {}).items():
        fetcher = FETCHERS.get(provider)
        if not fetcher:
            print(f"  ! unknown provider '{provider}', skipping")
            continue
        for entry in entries or []:
            if isinstance(entry, dict):
                slug = entry.get("slug")
                display = entry.get("name") or slug
            else:
                slug, display = entry, entry
            try:
                jobs = fetcher(slug)
            except Exception as exc:
                failed.append(f"{provider}/{slug}: {exc}")
                continue
            rows = [(
                j["job_key"], j["source"], display, norm_employer(display),
                j["title"], j["location"], j["url"], j["posted"],
                j["description"][:20000], time.strftime("%Y-%m-%d %H:%M"),
            ) for j in jobs]
            con.executemany(
                "INSERT OR REPLACE INTO jobs (job_key, source, company, company_norm, "
                "title, location, url, posted, description, fetched_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)", rows)
            con.commit()
            total += len(rows)
            print(f"  {display:<28} {provider:<11} {len(rows):>4} postings")
            time.sleep(args.delay)

    print(f"\n{total:,} postings stored.")
    if failed:
        print(f"{len(failed)} board(s) failed. Usually a wrong slug:")
        for f in failed:
            print("  -", f)


# ------------------------------------------------------------------- discover
#
# Turns the employer list you already loaded from DOL into a company list, by
# guessing job board slugs from employer names and probing which ones are real.

# Words dropped when building slug guesses, on top of the corporate suffixes in
# _SUFFIXES. These appear in legal names but rarely in board URLs.
_SLUG_DROP = {"technologies", "technology", "systems", "solutions", "services",
              "software", "labs", "laboratories", "international", "global",
              "worldwide", "enterprises", "industries", "partners", "ventures"}

PROBE_URLS = {
    "greenhouse": "https://boards-api.greenhouse.io/v1/boards/{slug}/jobs",
    "lever": "https://api.lever.co/v0/postings/{slug}?mode=json&limit=1",
    "ashby": "https://api.ashbyhq.com/posting-api/job-board/{slug}",
}


def slug_candidates(name):
    """Plausible board slugs for an employer name, most likely first.

    Deliberately conservative. It does not try single-token guesses on
    multi-word names, because 'definitive' from 'Definitive Intelligence' would
    happily match some unrelated company's board and quietly poison the results.
    """
    base = norm_employer(name)
    if not base:
        return []
    toks = [t for t in base.split() if t not in _SLUG_DROP] or base.split()
    joined = "".join(toks)
    out = [joined]
    if len(toks) > 1:
        out.append("-".join(toks))
    # Variants that keep the filler words ("benchlingtechnologies"). Generated
    # whenever anything was dropped, not just for multi-token names, or a name
    # that reduces to a single token would never get this guess.
    full = "".join(base.split())
    if full != joined:
        out.append(full)
        out.append("-".join(base.split()))
    seen, uniq = set(), []
    for s in out:
        if s and s not in seen and 2 <= len(s) <= 40:
            seen.add(s)
            uniq.append(s)
    return uniq


def probe(provider, slug, timeout=12):
    """Return (ok, n_jobs). A real board with zero openings still counts as ok."""
    url = PROBE_URLS[provider].format(slug=slug)
    try:
        r = requests.get(url, headers=UA, timeout=timeout)
    except requests.RequestException:
        return False, 0
    if r.status_code != 200:
        return False, 0
    try:
        data = r.json()
    except ValueError:
        return False, 0
    if provider == "greenhouse":
        return isinstance(data, dict) and "jobs" in data, len(data.get("jobs", []))
    if provider == "lever":
        return isinstance(data, list), len(data)
    return isinstance(data, dict) and "jobs" in data, len(data.get("jobs", []))


DEFAULT_ROLES = ("software", "developer", "engineer", "data scien", "data analyst",
                 "machine learning", "computer", "research", "programmer",
                 "statistician", "analyst")


def cmd_discover(args):
    con = connect()
    roles = [r.strip().lower() for r in args.roles.split(",") if r.strip()] \
        if args.roles else list(DEFAULT_ROLES)
    states = [s.strip().upper() for s in args.states.split(",") if s.strip()] \
        if args.states else None

    rows = con.execute(
        "SELECT employer_norm, employer_display, certified, titles, states "
        "FROM employers WHERE certified >= ? ORDER BY certified DESC",
        (args.min_certified,)).fetchall()
    if not rows:
        raise SystemExit("No employers loaded. Run `load-lca` first.")

    candidates = []
    for norm, display, certified, titles_json, states_json in rows:
        titles = " ".join(json.loads(titles_json or "[]")).lower()
        if roles and not any(r in titles for r in roles):
            continue
        if states:
            emp_states = json.loads(states_json or "[]")
            if not any(s in emp_states for s in states):
                continue
        # No name-based exclusion. Filtering employers by whether their name
        # contains "consulting" drops Deloitte and Palantir while keeping TCS,
        # Infosys and Accenture, which is arbitrary and silently loses good
        # employers. Volume and wage-level data are in the report instead, so
        # you can judge rather than have the tool judge for you.
        if args.max_certified and certified > args.max_certified:
            continue
        candidates.append((norm, display, certified))
        if len(candidates) >= args.limit:
            break

    print(f"{len(candidates):,} candidate employers "
          f"(certified >= {args.min_certified}"
          f"{', states ' + ','.join(states) if states else ', nationwide'}).")

    cache = {(p, s): (ok, n) for p, s, ok, n in con.execute(
        "SELECT provider, slug, ok, n_jobs FROM probe_cache")}

    tasks = []
    for norm, display, certified in candidates:
        for slug in slug_candidates(display):
            for provider in PROBE_URLS:
                if (provider, slug) in cache:
                    continue
                tasks.append((provider, slug, display, certified))

    print(f"{len(tasks):,} probes to run "
          f"({len(cache):,} already cached). Ctrl-C is safe, results are saved as they land.")

    found, done = {}, 0
    if tasks:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futs = {pool.submit(probe, p, s): (p, s, d, c) for p, s, d, c in tasks}
            try:
                for fut in as_completed(futs):
                    provider, slug, display, certified = futs[fut]
                    try:
                        ok, n = fut.result()
                    except Exception:
                        ok, n = False, 0
                    done += 1
                    con.execute(
                        "INSERT OR REPLACE INTO probe_cache VALUES (?,?,?,?,?)",
                        (provider, slug, int(ok), n, time.strftime("%Y-%m-%d")))
                    if ok:
                        cache[(provider, slug)] = (True, n)
                        print(f"  HIT {provider:<11} {slug:<28} {n:>4} jobs   ({display})")
                    if done % 300 == 0:
                        con.commit()
                        print(f"  ...{done:,}/{len(tasks):,}")
            except KeyboardInterrupt:
                print("\nInterrupted, saving what we have.")
            finally:
                con.commit()

    # Rebuild the company list from every cached hit that maps to a candidate.
    by_display = {}
    for norm, display, certified in candidates:
        for slug in slug_candidates(display):
            for provider in PROBE_URLS:
                ok, n = cache.get((provider, slug), (False, 0))
                if ok:
                    prev = by_display.get(display)
                    if prev is None or n > prev[2]:
                        by_display[display] = (provider, slug, n, certified)

    existing = {}
    if args.merge and os.path.exists(args.out):
        with open(args.out) as fh:
            existing = (yaml.safe_load(fh) or {}).get("companies") or {}

    merged = {p: list(existing.get(p) or []) for p in PROBE_URLS}
    seen = {p: {(e.get("slug") if isinstance(e, dict) else e) for e in merged[p]}
            for p in PROBE_URLS}
    added = 0
    for display, (provider, slug, n, certified) in sorted(
            by_display.items(), key=lambda kv: -kv[1][3]):
        if slug in seen[provider]:
            continue
        merged[provider].append({"slug": slug, "name": display})
        seen[provider].add(slug)
        added += 1

    with open(args.out, "w") as fh:
        fh.write("# Generated by `sponsorscan.py discover`. Hand edits to `name:`\n"
                 "# are preserved on re-run with --merge (the default).\n"
                 "#\n"
                 "# Slugs are guessed from DOL employer names and confirmed by a live\n"
                 "# probe, so a listed board definitely exists. It is still possible for\n"
                 "# a guess to land on a DIFFERENT company with a similar name. If a\n"
                 "# company's postings look wrong, delete its line.\n\n")
        yaml.safe_dump({"companies": {p: merged[p] for p in PROBE_URLS if merged[p]}},
                       fh, sort_keys=False, default_flow_style=False)

    total = sum(len(v) for v in merged.values())
    print(f"\n{len(by_display):,} employers matched to a live board. "
          f"Added {added:,} new; {total:,} companies now in {args.out}.")
    print("Next:  python sponsorscan.py fetch-jobs --replace")


# ---------------------------------------------------------------------- report

# Phrases that mean you are excluded regardless of anything else. Checked against
# the posting body. This is the filter that actually matters while on OPT.
DISQUALIFIERS = [
    # Any negation followed by "sponsor" within the same sentence. Catches the
    # long tail of phrasings ("do not offer", "does not provide", "unable to",
    # "without", "no sponsorship available", "not now or in the future require")
    # without needing a pattern per variant. Bounded by [^.] so it cannot leak
    # across a sentence boundary and match an unrelated negation.
    r"\b(?:not|no|non|unable|without|cannot|can't|won't|unwilling)\b[^.]{0,60}?sponsor",
    r"\bmust be (?:a |an )?(?:u\.?\s?s\.?|united states)\s?(?:citizen|person|national)\b",
    r"\b(?:u\.?\s?s\.?|united states)\s?citizenship (?:is )?required\b",
    r"\bsecurity clearance\b",
    r"\bitar\b",
    r"\bexport control(?:led|s)?\b",
]

SPONSOR_POSITIVE = [
    r"\bvisa sponsorship (?:is )?available\b",
    r"\bwe (?:will )?sponsor\b",
    r"\bh-?1b sponsorship\b",
    r"\bsponsorship (?:is )?(?:available|offered|provided)\b",
]

SENIOR_TITLE = re.compile(
    r"\b(senior|sr\.?|staff|principal|distinguished|lead|architect|manager|"
    r"director|head of|vp|vice president|chief|fellow)\b", re.I)

ENTRY_TITLE = re.compile(
    r"\b(intern|internship|new ?grad|new graduate|university grad|recent grad|"
    r"early career|entry.level|junior|jr\.?|associate|apprentice|i{1,2}\b)\b", re.I)

DISQ_RE = [re.compile(p, re.I) for p in DISQUALIFIERS]
POS_RE = [re.compile(p, re.I) for p in SPONSOR_POSITIVE]


def load_employer_index(con):
    rows = con.execute(
        "SELECT employer_norm, employer_display, certified, denied, withdrawn, titles, "
        "       states, lvl1, lvl2, lvl3, lvl4 FROM employers").fetchall()
    out = {}
    for r in rows:
        lvls = [r[7] or 0, r[8] or 0, r[9] or 0, r[10] or 0]
        tot = sum(lvls)
        filed = (r[2] or 0) + (r[3] or 0) + (r[4] or 0)
        out[r[0]] = {
            "display": r[1], "certified": r[2], "denied": r[3], "withdrawn": r[4],
            "titles": json.loads(r[5] or "[]"), "states": json.loads(r[6] or "[]"),
            "lvls": lvls,
            # Share filed at Level III/IV. Higher is better on two counts: the
            # pay is higher, and the FY2027 weighted lottery favours those levels.
            "senior_share": round((lvls[2] + lvls[3]) / tot, 2) if tot else None,
            "trouble_rate": round(((r[3] or 0) + (r[4] or 0)) / filed, 2) if filed else None,
        }
    return out


def match_employer(company_norm, index, keys, cutoff):
    """Exact key match, then fuzzy fallback."""
    if company_norm in index:
        return company_norm, 100
    if not HAVE_RAPIDFUZZ or not keys:
        return None, 0
    hit = rf_process.extractOne(
        company_norm, keys, scorer=rf_fuzz.token_set_ratio, score_cutoff=cutoff)
    if hit:
        return hit[0], int(hit[1])
    return None, 0


def cmd_report(args):
    con = connect()
    index = load_employer_index(con)
    keys = list(index.keys())
    if not index:
        raise SystemExit("No LCA data loaded. Run `load-lca` first.")

    locs = [s.strip().lower() for s in (args.locations or "").split(",") if s.strip()]
    results, dropped = [], 0

    for row in con.execute(
            "SELECT company, company_norm, title, location, url, posted, description, source "
            "FROM jobs"):
        company, cnorm, title, location, url, posted, desc, source = row
        blob = f"{title}\n{desc}"

        hit = next((p.pattern for p in DISQ_RE if p.search(blob)), None)
        if hit:
            dropped += 1
            continue

        if not args.include_senior and SENIOR_TITLE.search(title or ""):
            continue

        score, why = 0, []

        key, conf = match_employer(cnorm, index, keys, args.fuzzy_cutoff)
        emp = index.get(key) if key else None
        if emp and emp["certified"] > 0:
            score += 40
            why.append(f"{emp['certified']} certified LCAs")
            if conf < 100:
                why.append(f"fuzzy match {conf} to '{emp['display']}'")
            if emp["certified"] >= 25:
                score += 10
            if emp["senior_share"] is not None and emp["senior_share"] >= 0.5:
                score += 10
                why.append(f"{int(emp['senior_share']*100)}% filed at wage level III/IV")
            if emp["trouble_rate"] is not None and emp["trouble_rate"] >= 0.25:
                score -= 10
                why.append(f"{int(emp['trouble_rate']*100)}% denied or withdrawn")
            tl = (title or "").lower()
            if any(any(w in t for w in tl.split() if len(w) > 4) for t in emp["titles"]):
                score += 15
                why.append("filed for similar roles")
        elif args.sponsors_only:
            continue

        if ENTRY_TITLE.search(title or ""):
            score += 20
            why.append("entry level")

        if locs and any(l in (location or "").lower() for l in locs):
            score += 15
            why.append("location match")
        elif locs and not args.any_location:
            continue

        if any(p.search(blob) for p in POS_RE):
            score += 10
            why.append("sponsorship stated in posting")

        results.append({
            "score": score, "company": company, "title": title,
            "location": location, "posted": posted, "source": source,
            "signals": "; ".join(why), "url": url,
            "lca_certified": emp["certified"] if emp else 0,
            "lca_senior_wage_share": emp["senior_share"] if emp else None,
            "lca_denied_withdrawn_rate": emp["trouble_rate"] if emp else None,
        })

    results.sort(key=lambda r: (-r["score"], r["company"]))

    if args.out:
        with open(args.out, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(results[0].keys()) if results else
                               ["score", "company", "title", "location", "posted",
                                "source", "signals", "url", "lca_certified",
                                "lca_senior_wage_share", "lca_denied_withdrawn_rate"])
            w.writeheader()
            w.writerows(results)
        print(f"Wrote {len(results)} rows to {args.out}")

    print(f"\n{dropped} postings dropped on explicit disqualifiers "
          f"(citizenship, clearance, or no-sponsorship language).\n")
    for r in results[:args.top]:
        print(f"[{r['score']:>3}] {r['company']} - {r['title']}")
        print(f"      {r['location'] or '?'} | {r['posted'] or '?'} | {r['signals']}")
        print(f"      {r['url']}\n")


# ------------------------------------------------------------------------ main

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("load-lca", help="Load DOL LCA disclosure file into SQLite")
    a.add_argument("path", help="Local .xlsx/.csv path, or an https:// URL")
    a.add_argument("--replace", action="store_true", help="Clear existing employer rows first")
    a.set_defaults(func=cmd_load_lca)

    b = sub.add_parser("fetch-jobs", help="Pull live postings from ATS job boards")
    b.add_argument("--companies", default="companies.yaml")
    b.add_argument("--replace", action="store_true")
    b.add_argument("--delay", type=float, default=0.4, help="Seconds between boards")
    b.set_defaults(func=cmd_fetch_jobs)

    d = sub.add_parser("discover", help="Build companies.yaml from the DOL employer list")
    d.add_argument("--out", default="companies.yaml")
    d.add_argument("--min-certified", type=int, default=5,
                   help="Skip employers with fewer certified LCAs (default 5)")
    d.add_argument("--roles", default="",
                   help="Comma-separated title keywords; blank uses a tech/data default set")
    d.add_argument("--states", default="",
                   help="Comma-separated 2-letter states; blank means nationwide")
    d.add_argument("--limit", type=int, default=2000, help="Max employers to probe")
    d.add_argument("--workers", type=int, default=12)
    d.add_argument("--max-certified", type=int, default=0,
                   help="Skip employers above this many certified LCAs "
                        "(0 = no cap). Use to exclude the handful of mega-filers.")
    d.add_argument("--no-merge", dest="merge", action="store_false",
                   help="Overwrite the company list instead of merging into it")
    d.set_defaults(func=cmd_discover, merge=True)

    c = sub.add_parser("report", help="Join, filter, rank, export")
    c.add_argument("--out", default="matches.csv")
    c.add_argument("--top", type=int, default=25)
    c.add_argument("--locations", default="",
                   help="Comma-separated location substrings; blank means anywhere in the US")
    c.add_argument("--any-location", action="store_true",
                   help="Keep postings that miss the location filter")
    c.add_argument("--include-senior", action="store_true")
    c.add_argument("--sponsors-only", action="store_true",
                   help="Drop employers with no certified LCAs on record")
    c.add_argument("--fuzzy-cutoff", type=int, default=90)
    c.set_defaults(func=cmd_report)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
